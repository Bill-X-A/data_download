import requests
import lark_oapi as lark
from lark_oapi.api.bitable.v1 import *
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import io
import streamlit as st

APP_ID = st.secrets["FEISHU_APP_ID"]
APP_SECRET = st.secrets["FEISHU_APP_SECRET"]
APP_TOKEN = "AO2NbKrqNaWFZ9suHKJcjGYLn4b"
TABLE_ID = "tbl1W59w24xZBvNc"

def get_from_feishu():
    client = lark.Client.builder() \
        .app_id(APP_ID) \
        .app_secret(APP_SECRET) \
        .build()

    request = ListAppTableRecordRequest.builder() \
        .app_token(APP_TOKEN) \
        .table_id(TABLE_ID) \
        .build()

    response = client.bitable.v1.app_table_record.list(request)
    
    if response.success():
        records = []
        for item in response.data.items:
            records.append(item.fields)
        return records
    else:
        return []

def export_with_images():
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    
    token = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": APP_ID, "app_secret": APP_SECRET}
    ).json()["tenant_access_token"]
    
    records = get_from_feishu()
    
    wb = Workbook()
    ws = wb.active
    ws.append(["销售员", "手机号", "商品名称", "SN码", "订单金额", "购买品类", "购买用途", "会员等级", "时间", "三码合一照片"])
    
    row = 2
    for fields in records:
        ws.cell(row=row, column=1, value=str(fields.get("销售员", "")))
        ws.cell(row=row, column=2, value=str(fields.get("手机号", "")))
        ws.cell(row=row, column=3, value=str(fields.get("商品名称", "")))
        ws.cell(row=row, column=4, value=str(fields.get("SN码", "")))
        ws.cell(row=row, column=5, value=str(fields.get("订单金额", "")))
        ws.cell(row=row, column=6, value=str(fields.get("购买品类", "")))
        ws.cell(row=row, column=7, value=str(fields.get("购买用途", "")))
        ws.cell(row=row, column=8, value=str(fields.get("会员等级", "")))
        ws.cell(row=row, column=9, value=str(fields.get("时间", "")))
        
        photo = fields.get("三码合一照片")
        if photo and isinstance(photo, list):
            file_token = photo[0].get("file_token")
            if file_token:
                url = f"https://open.feishu.cn/open-apis/drive/v1/medias/{file_token}/download"
                img_data = requests.get(url, headers={"Authorization": f"Bearer {token}"}).content
                img = XLImage(io.BytesIO(img_data))
                img.width = 100
                img.height = 100
                ws.row_dimensions[row].height = 80
                ws.add_image(img, f"J{row}")
        row += 1
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

st.divider()
if st.button("导出带图记录"):
    with st.spinner("正在生成，请稍候..."):
        excel_data = export_with_images()
    st.download_button(
        label="📥 下载Excel",
        data=excel_data,
        file_name="销售记录带图.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
